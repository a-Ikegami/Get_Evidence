import os
import json

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class MergeEvi:

    book: Workbook
    config: any

    def __init__(self) -> None:
        try:
            with open("config.json") as config_file:
                config = json.load(config_file)

                self.book = Workbook()
                self.config = config
        except FileExistsError as e:
            print(e)
            return

    def input_excel(self, ws: Worksheet, evi_list: list):
        """
        ws(シート名)毎に対象のエビデンスファイル(evi_list)をマージする
        """

        for ename in evi_list:
            evi_path = self.get_evi_path(ename)
            try:
                with open(evi_path, "r", encoding="utf-8") as f:
                    max = ws.max_row
                    for y, l in enumerate(f, start=1):
                        ws.cell(row=max + y, column=1).value = l
            except FileNotFoundError:
                pass

    def get_evi_path(self, evi_name: str) -> str:
        """対象のエビデンスファイルパスを生成します
        Args:
            evi_name(string)
        """

        return "evi/" + evi_name + ".txt"

    def exist_evi(self, evi_list: list) -> bool:
        """対象のエビデンスリストにいずれか1つでも存在するかを返します
        Args:
            evi_list(list) エビデンスリスト

        Returns:
            bool: True=存在する/False=存在しない
        """

        exist: bool = False

        for evi_name in evi_list:
            path = self.get_evi_path(evi_name)

            if os.path.exists(path):
                exist = True

        return exist

    def create_sheet(self, sheet_name: str, evi_list: list):
        """対象のエビデンスが存在する場合、シートを作成する

        存在しない場合はFalseを返す。

        """

        if self.exist_evi(evi_list):
            self.book.create_sheet(sheet_name)
        else:
            return False

    def main(self):
        output_path = self.config["setting"]["output_path"]

        # 出力ファイルが存在する場合、削除する。
        if os.path.exists(output_path):
            os.remove(output_path)

        # エビデンスファイルをエクセルに書き込む。
        for evidence in self.config["evidence"]:
            sheet_name = evidence["sheet_name"]

            # シート作成を行わなかった場合はエビデンスが無いので次へ。
            if self.create_sheet(sheet_name, evidence["file_list"]) == False:
                continue

            work_sheet = self.book[sheet_name]
            self.input_excel(work_sheet, evidence["file_list"])

        # エクセルを保存する。
        self.book.save(output_path)


if __name__ == "__main__":
    instance = MergeEvi()
    instance.main()
